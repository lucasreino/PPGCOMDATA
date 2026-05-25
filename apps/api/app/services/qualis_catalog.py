"""Carrega e consulta catálogo Qualis (ISSN / título → estrato)."""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl


def normalize_issn(value: Optional[str]) -> str:
    if not value:
        return ""
    raw = str(value).upper().strip()
    digits = re.sub(r"[^0-9X]", "", raw)
    if len(digits) >= 8:
        return digits[:8]
    if len(digits) >= 4:
        return digits
    return ""


def normalize_title(value: Optional[str]) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFD", str(value).strip().upper())
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^A-Z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    for stop in ("REVISTA DE ", "REVISTA ", "JOURNAL OF ", "JOURNAL "):
        if text.startswith(stop):
            text = text[len(stop) :].strip()
    return text


def normalize_estrato(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    e = str(value).strip().upper()
    e = e.replace("ESTRATO", "").strip()
    if re.match(r"^[A-C][1-5]?$", e):
        return e
    return None


@dataclass
class QualisEntry:
    issn: str
    titulo: str
    estrato: str
    issn_key: str
    titulo_key: str


def default_manual_overrides_path() -> Path:
    here = Path(__file__).resolve()
    candidates: list[Path] = [
        Path("/workspace/data/qualis/manual_overrides.json"),
    ]
    if len(here.parents) > 2:
        candidates.append(here.parents[2] / "data" / "qualis" / "manual_overrides.json")
    if len(here.parents) > 4:
        candidates.append(here.parents[4] / "data" / "qualis" / "manual_overrides.json")
    candidates.append(Path.cwd() / "data" / "qualis" / "manual_overrides.json")
    for p in candidates:
        if p.is_file():
            return p
    return candidates[0]


def load_manual_overrides(path: Optional[Path] = None) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Retorna (by_issn, by_veiculo) com estratos validados."""
    p = path or default_manual_overrides_path()
    if not p.is_file():
        return {}, {}
    raw = json.loads(p.read_text(encoding="utf-8"))
    by_issn: Dict[str, str] = {}
    by_veiculo: Dict[str, str] = {}
    for key, val in (raw.get("by_issn") or {}).items():
        issn_key = normalize_issn(key)
        estrato = normalize_estrato(val)
        if issn_key and estrato:
            by_issn[issn_key] = estrato
    for key, val in (raw.get("by_veiculo") or {}).items():
        titulo_key = normalize_title(key)
        estrato = normalize_estrato(val)
        if titulo_key and estrato:
            by_veiculo[titulo_key] = estrato
    return by_issn, by_veiculo


def default_qualis_xlsx_path() -> Path:
    here = Path(__file__).resolve()
    candidates: list[Path] = [
        Path("/workspace/data/qualis/qualis-comunicacao-2026.xlsx"),
    ]
    if len(here.parents) > 2:
        candidates.append(
            here.parents[2] / "data" / "qualis" / "qualis-comunicacao-2026.xlsx"
        )
    if len(here.parents) > 4:
        candidates.append(
            here.parents[4] / "data" / "qualis" / "qualis-comunicacao-2026.xlsx"
        )
    candidates.append(Path.cwd() / "data" / "qualis" / "qualis-comunicacao-2026.xlsx")

    for p in candidates:
        if p.is_file():
            return p
    return candidates[0]


def load_qualis_catalog(xlsx_path: Path) -> Tuple[List[QualisEntry], Dict[str, str], Dict[str, str]]:
    """
    Retorna (entries, by_issn, by_titulo).
    by_titulo: último estrato se houver duplicata de título normalizado.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    entries: List[QualisEntry] = []
    by_issn: Dict[str, str] = {}
    by_titulo: Dict[str, str] = {}

    max_row = sheet.max_row or 0
    for row_idx in range(2, max_row + 1):
        row = (
            sheet.cell(row_idx, 1).value,
            sheet.cell(row_idx, 2).value,
            sheet.cell(row_idx, 3).value,
        )
        if not row or not any(row):
            continue
        issn_raw = row[0] if len(row) > 0 else None
        titulo_raw = row[1] if len(row) > 1 else None
        estrato_raw = row[2] if len(row) > 2 else None
        estrato = normalize_estrato(estrato_raw)
        if not estrato:
            continue
        issn_key = normalize_issn(issn_raw)
        titulo_key = normalize_title(titulo_raw)
        if not issn_key and not titulo_key:
            continue
        entry = QualisEntry(
            issn=str(issn_raw or "").strip(),
            titulo=str(titulo_raw or "").strip(),
            estrato=estrato,
            issn_key=issn_key,
            titulo_key=titulo_key,
        )
        entries.append(entry)
        if issn_key:
            by_issn[issn_key] = estrato
        if titulo_key:
            by_titulo[titulo_key] = estrato

    wb.close()
    return entries, by_issn, by_titulo


def lookup_qualis(
    issn: Optional[str],
    veiculo: Optional[str],
    by_issn: Dict[str, str],
    by_titulo: Dict[str, str],
    manual_issn: Optional[Dict[str, str]] = None,
    manual_veiculo: Optional[Dict[str, str]] = None,
) -> Tuple[Optional[str], str]:
    """Retorna (estrato, metodo)."""
    issn_key = normalize_issn(issn)
    veic_key = normalize_title(veiculo)

    if manual_issn and issn_key and issn_key in manual_issn:
        return manual_issn[issn_key], "manual_issn"
    if manual_veiculo and veic_key and veic_key in manual_veiculo:
        return manual_veiculo[veic_key], "manual_veiculo"

    if issn_key and issn_key in by_issn:
        return by_issn[issn_key], "issn"

    if not veic_key:
        return None, "none"

    if veic_key in by_titulo:
        return by_titulo[veic_key], "titulo_exato"

    for cat_key, estrato in by_titulo.items():
        if len(cat_key) < 8 or len(veic_key) < 8:
            continue
        if cat_key in veic_key or veic_key in cat_key:
            return estrato, "titulo_parcial"

    return None, "none"


def apply_qualis_to_producoes(
    session,
    *,
    tipos: frozenset[str] | None = None,
    xlsx_path: Path | None = None,
) -> dict[str, int]:
    """
    Cruza artigos/anais com o catálogo Qualis e grava estrato em Producao.qualis.
    """
    from collections import Counter
    from sqlmodel import select

    from app.models.data import Producao

    path = xlsx_path or default_qualis_xlsx_path()
    if not path.is_file():
        return {"atualizado": 0, "ja_ok": 0, "sem_match": 0}

    _, by_issn, by_titulo = load_qualis_catalog(path)
    manual_issn, manual_veiculo = load_manual_overrides()
    allowed = tipos or frozenset({"artigo", "anais"})
    stats: Counter = Counter()

    for prod in session.exec(select(Producao)).all():
        if (prod.tipo or "").lower() not in allowed:
            continue
        estrato, _ = lookup_qualis(
            prod.issn,
            prod.veiculo,
            by_issn,
            by_titulo,
            manual_issn,
            manual_veiculo,
        )
        if not estrato:
            stats["sem_match"] += 1
            continue
        if normalize_estrato(prod.qualis) == estrato:
            stats["ja_ok"] += 1
            continue
        prod.qualis = estrato
        session.add(prod)
        stats["atualizado"] += 1

    if stats["atualizado"]:
        session.commit()

    return {
        "atualizado": stats["atualizado"],
        "ja_ok": stats["ja_ok"],
        "sem_match": stats["sem_match"],
    }
